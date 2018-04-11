#!usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import worksheet
from collections import OrderedDict
import datetime

def combine(file):
    
    wb = load_workbook(filename=file)
    sheet_students = wb['students']
    sheet_time = wb['time']
    sheet_combine = wb.create_sheet('combine', 2)
    worksheet.copier.WorksheetCopy(sheet_students, sheet_combine).copy_worksheet()
    course_attended_time = {}
    for row in range(2, sheet_time.max_row + 1):
        course_attended_time[sheet_time['B'+str(row)].value] = sheet_time['C'+str(row)].value
    sheet_combine['D1'] = '学习时间'
    for row in range(2, sheet_combine.max_row + 1):
        sheet_combine['D'+str(row)] = course_attended_time[sheet_combine['B'+str(row)].value]
    wb.save(filename=file)
    
       

def split(file):
    wb = load_workbook(filename=file)
    sheet_combine = wb['combine']
    combine_dic = {}
    for row in range(2, sheet_combine.max_row + 1):
        combine_dic[sheet_combine.cell(row=row, column=2).value] = []  #strftime('%Y/%m/%d %H:%M')
        for col in range(1, sheet_combine.max_column + 1):
            combine_dic[sheet_combine.cell(row=row, column=2).value].append(sheet_combine.cell(row=row, column=col).value)
    combine_dic_ordered = OrderedDict(sorted(combine_dic.items(), key=lambda t: t[1][0]))
    split_marker = list(combine_dic_ordered.values())[0][0].strftime('%Y')
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = split_marker
    ws1['A1'] = '创建时间'
    ws1['B1'] = '课程名称'
    ws1['C1'] = '学习人数'
    ws1['D1'] = '学习时间'
    ws1['A2'] = list(combine_dic_ordered.values())[0][0]#.strftime('%Y/%m/%d %H:%M')
    ws1['B2'] = list(combine_dic_ordered.values())[0][1]
    ws1['C2'] = list(combine_dic_ordered.values())[0][2]
    ws1['D2'] = list(combine_dic_ordered.values())[0][3]
    i = 3
    for key, value in combine_dic_ordered.items():
        if value[0].strftime('%Y') == split_marker:
            for j in range(1,5):
                ws1.cell(row=i, column=j).value = value[j-1]
        else: #value[0].strftime('%Y') != split_marker:
            wb1.save(str(split_marker)+'.'+'xlsx')
            wb1.close()
            split_marker = value[0].strftime('%Y')
            wb1 = Workbook()
            ws1 = wb1.active
            ws1.title = split_marker
            ws1['A1'] = '创建时间'
            ws1['B1'] = '课程名称'
            ws1['C1'] = '学习人数'
            ws1['D1'] = '学习时间'
            ws1['A2'] = value[0]#.strftime('%Y/%m/%d %H:%M')
            ws1['B2'] = value[1]
            ws1['C2'] = value[2]
            ws1['D2'] = value[3]
            i = 2
        i += 1
    wb1.save(str(split_marker)+'.'+'xlsx')
    wb1.close()    
    
            
if __name__ == '__main__':
    file = 'courses.xlsx'
    combine(file)
    split(file)
